from django.db.models import Q
from .models import Order, CompanyUser
import csv
import openpyxl
from django.http import HttpResponse

def export_orders_to_excel(orders):
    """Экспорт заказов в Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заказы"
    
    # Заголовки
    headers = ['ID', 'Клиент', 'Статус', 'Сумма', 'Дата создания']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Данные
    for row, order in enumerate(orders, 2):
        ws.cell(row=row, column=1, value=order.id)
        ws.cell(row=row, column=2, value=order.client.name)
        ws.cell(row=row, column=3, value=order.get_status_display())
        ws.cell(row=row, column=4, value=float(order.total_amount))
        ws.cell(row=row, column=5, value=order.created_at.date())
    
    # HTTP ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=orders.xlsx'
    wb.save(response)
    return response

def get_user_orders(user, company):
    """Получить заказы пользователя в зависимости от его роли"""
    try:
        company_user = CompanyUser.objects.get(user=user, company=company)
        
        if company_user.role in ['owner', 'manager']:
            # Владельцы и менеджеры видят все заказы компании
            return Order.objects.filter(company=company)
        elif company_user.role == 'executor':
            # Исполнители видят только назначенные им заказы
            return Order.objects.filter(
                company=company,
                executors__user=user
            ).distinct()
        else:
            return Order.objects.none()
    except CompanyUser.DoesNotExist:
        return Order.objects.none()


def can_user_access_order(user, order):
    """Проверить, может ли пользователь получить доступ к заказу"""
    try:
        company_user = CompanyUser.objects.get(
            user=user, 
            company=order.company,
            is_active=True
        )
        
        if company_user.role in ['owner', 'manager']:
            return True
        elif company_user.role == 'executor':
            return order.executors.filter(user=user).exists()
        
        return False
    except CompanyUser.DoesNotExist:
        return False


def calculate_user_commission(user, company, period_year, period_month):
    """Рассчитать комиссию пользователя за период"""
    try:
        company_user = CompanyUser.objects.get(user=user, company=company)
        
        # Получаем завершенные заказы пользователя за период
        completed_orders = Order.objects.filter(
            company=company,
            status='completed',
            completion_date__year=period_year,
            completion_date__month=period_month,
            executors__user=user
        ).distinct()
        
        total_commission = 0
        for order in completed_orders:
            # Рассчитываем комиссию как процент от суммы заказа
            commission = (order.total_amount * company_user.commission_rate) / 100
            total_commission += commission
        
        return total_commission
    except CompanyUser.DoesNotExist:
        return 0